#!/usr/bin/env python3
"""Build a friendly Excel report from the edge CSV log.

Workbook layout:
  - "Overview"       : headline numbers + a per-person totals table
  - "Daily (all)"    : every (date, person) row across all days
  - one sheet PER PERSON : that person's day-by-day history (date, time, year,
                           first/last seen, working/phone/idle minutes, productivity)

Daily data accumulates in events_log.csv (it is never wiped), so this report grows
day by day automatically. Importable as build_report() (used by the web app's
Download button) or runnable directly (export_excel.bat).

Usage:
  python export_excel.py [events_log.csv] [AI_Office_Report.xlsx]
"""
from __future__ import annotations
import sys
import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FONT = "Arial"
NAVY = "1F3864"
BLUE = "2F5597"
BAND = "F5F8FD"
GREY = "595959"
_THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
WEIGHTS = {"working": 1.0, "meeting": 0.9, "talking": 0.5, "walking": 0.3, "idle": 0.0, "phone": -0.2}


def _set(cell, value, *, size=10, bold=False, color="000000", fill=None, align="left", border=True):
    cell.value = value
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if border:
        cell.border = BORDER


def _table(ws, top, headers, rows):
    for j, h in enumerate(headers, start=1):
        _set(ws.cell(row=top, column=j), h, size=11, bold=True, color="FFFFFF", fill=BLUE, align="center")
    for r, row in enumerate(rows, start=top + 1):
        band = BAND if (r - top) % 2 == 0 else "FFFFFF"
        for j, val in enumerate(row, start=1):
            _set(ws.cell(row=r, column=j), val, fill=band, align="center" if j > 1 else "left")
    widths = []
    for j, h in enumerate(headers):
        col = [str(h)] + [str(rw[j]) for rw in rows]
        widths.append(min(34, max(12, max(len(x) for x in col) + 2)))
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=top + 1, column=1)


def _stats(g):
    a = g["activity"].fillna("").astype(str).value_counts()
    w = int(a.get("working", 0)); idle = int(a.get("idle", 0)); ph = int(a.get("phone", 0))
    wk = int(a.get("walking", 0)); mt = int(a.get("meeting", 0))
    active = max(1, w + idle + ph + wk + mt)
    weighted = w * WEIGHTS["working"] + mt * WEIGHTS["meeting"] + wk * WEIGHTS["walking"] + ph * WEIGHTS["phone"]
    score = round(max(0.0, min(100.0, 100.0 * weighted / active)), 1)
    present = round((g["timestamp"].max() - g["timestamp"].min()).total_seconds() / 60, 1)
    return {"first": g["timestamp"].min().strftime("%I:%M %p"),
            "last": g["timestamp"].max().strftime("%I:%M %p"),
            "present": present, "working": round(w / 60, 1), "phone": round(ph / 60, 1),
            "idle": round(idle / 60, 1), "walking": round(wk / 60, 1), "score": score}


def _sheet_name(name, used):
    s = re.sub(r"[\\/?*\[\]:]", "", str(name)).strip()[:28] or "Person"
    base, i = s, 2
    while s.lower() in used:
        s = f"{base[:25]}_{i}"
        i += 1
    used.add(s.lower())
    return s


def _empty(out_path, msg):
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    _set(ws["A1"], "AI Office - Activity Report", size=18, bold=True, color=NAVY, border=False)
    _set(ws["A3"], msg, size=11, color=GREY, border=False)
    ws.column_dimensions["A"].width = 70
    wb.save(out_path)


def build_report(csv_path="events_log.csv", out_path="AI_Office_Report.xlsx"):
    if not os.path.exists(csv_path) or os.path.getsize(csv_path) == 0:
        _empty(out_path, "Abhi koi data nahi. Camera thodi der chalao, phir dobara download karo.")
        return out_path
    df = pd.read_csv(csv_path)
    if df.empty:
        _empty(out_path, "Abhi koi data nahi (file khaali hai).")
        return out_path
    # Old rows may carry a tz suffix (+05:30 / Z) and new ones not -> that "mixed
    # timezones" mix crashes pandas. Strip any tz suffix (keep the wall-clock time).
    _raw = df["timestamp"].astype(str).str.replace(r'([+-]\d{2}:?\d{2}|Z)\s*$', '', regex=True)
    df["timestamp"] = pd.to_datetime(_raw, errors="coerce", format="mixed")
    df = df.dropna(subset=["timestamp"]).sort_values("timestamp")
    if df.empty:
        _empty(out_path, "Timestamps padh nahi paaye.")
        return out_path

    df["DateKey"] = df["timestamp"].dt.date
    df["DateStr"] = df["timestamp"].dt.strftime("%d %b %Y")
    df["Name"] = df["subject_id"].apply(
        lambda s: "Unknown" if (pd.isna(s) or str(s).strip().lower() in ("", "unknown")) else str(s))
    beh = df[df["type"] == "behavior"]
    if beh.empty:
        beh = df

    # per (date, person) aggregation
    daily = []
    for (dk, name), g in beh.groupby(["DateKey", "Name"]):
        st = _stats(g)
        daily.append({"DateKey": dk, "Date": g["DateStr"].iloc[0],
                      "Day": dk.strftime("%A"), "Name": name, **st})
    daily.sort(key=lambda r: (r["DateKey"], r["Name"]), reverse=True)

    wb = Workbook()

    # ---- Overview ----
    ov = wb.active
    ov.title = "Overview"
    ov.sheet_view.showGridLines = False
    ov.merge_cells("A1:H1")
    _set(ov["A1"], "AI Office  -  Activity Report", size=20, bold=True, color="FFFFFF", fill=NAVY, align="center", border=False)
    ov.row_dimensions[1].height = 32
    gen = df["timestamp"].max().strftime("%d %b %Y, %I:%M %p")
    rng = f"{df['timestamp'].min().strftime('%d %b %Y')}  to  {df['timestamp'].max().strftime('%d %b %Y')}"
    ov.merge_cells("A2:H2")
    _set(ov["A2"], f"Generated: {gen}     Data range: {rng}", size=11, color=GREY, align="center", border=False)
    people = sorted(beh["Name"].unique())
    info = [("People recorded", len(people)), ("Days recorded", beh["DateKey"].nunique()),
            ("Total detections", len(df))]
    r = 4
    for label, val in info:
        _set(ov.cell(row=r, column=1), label, bold=True)
        _set(ov.cell(row=r, column=2), val)
        r += 1
    r += 1
    _set(ov.cell(row=r, column=1), "Per person (totals)", size=12, bold=True, color="FFFFFF", fill=BLUE, border=False)
    ov.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    prows = []
    for name in people:
        gp = [d for d in daily if d["Name"] == name]
        days = len({d["DateKey"] for d in gp})
        tot = round(sum(d["present"] for d in gp), 1)
        avg = round(sum(d["score"] for d in gp) / len(gp), 1) if gp else 0
        prows.append([name, days, tot, avg])
    _table(ov, r, ["Person", "Days seen", "Total time (min)", "Avg productivity %"], prows)
    if prows:
        ch = BarChart(); ch.type = "col"; ch.title = "Avg productivity per person"; ch.legend = None
        ch.height = 7; ch.width = 13
        data = Reference(ov, min_col=4, min_row=r, max_row=r + len(prows))
        cats = Reference(ov, min_col=1, min_row=r + 1, max_row=r + len(prows))
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ov.add_chart(ch, "F3")

    # ---- Daily (all) ----
    da = wb.create_sheet("Daily (all)")
    da.sheet_view.showGridLines = False
    drows = [[d["Date"], d["Day"], d["Name"], d["first"], d["last"], d["present"], d["working"], d["phone"], d["idle"], d["score"]]
             for d in daily]
    _table(da, 1, ["Date", "Day", "Person", "First seen", "Last seen", "Time present (min)",
                   "Working (min)", "Phone (min)", "Idle (min)", "Productivity %"], drows)

    # ---- one sheet per person ----
    used = {"overview", "daily (all)"}
    for name in people:
        gp = [d for d in daily if d["Name"] == name]
        rows = [[d["Date"], d["Day"], d["first"], d["last"], d["present"], d["working"], d["phone"], d["idle"], d["walking"], d["score"]]
                for d in gp]
        ws = wb.create_sheet(_sheet_name(name, used))
        ws.sheet_view.showGridLines = False
        _set(ws["A1"], f"{name}  -  day-by-day", size=14, bold=True, color=NAVY, border=False)
        ws.merge_cells("A1:J1")
        _table(ws, 3, ["Date", "Day", "First seen", "Last seen", "Time present (min)", "Working (min)",
                       "Phone (min)", "Idle (min)", "Walking (min)", "Productivity %"], rows)
        if rows:
            ch = BarChart(); ch.type = "col"; ch.title = "Working vs Phone (min) per day"
            ch.height = 7; ch.width = 14
            data = Reference(ws, min_col=6, max_col=7, min_row=3, max_row=3 + len(rows))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(rows))
            ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
            ws.add_chart(ch, f"A{3 + len(rows) + 2}")

    wb.save(out_path)
    return out_path


def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else "events_log.csv"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "AI_Office_Report.xlsx"
    p = build_report(csv_path, out_path)
    print(f"Wrote '{p}'.")


if __name__ == "__main__":
    main()
